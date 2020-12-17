import unittest
from openpyxl import load_workbook, cell
import openpyxl

class Test(unittest.TestCase):

    def test_read_excel_column_file(self):

        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("DummySheet", 1)
        sheet["A1"].value = "This is sample writing"
        wb.save("test-excel.xlsx")

if __name__ == "__main__":
    unittest.main()