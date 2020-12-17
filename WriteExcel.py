from os import startfile
from openpyxl import load_workbook    

workbook = load_workbook("Test.xlsx")
sheet = workbook.active

# sheet.cell(row = 1, column = 1, value = "Hello")
sheet["A1"] = "Hello"

workbook.save("Test.xlsx")

# now file will start
startfile("Test.xls")